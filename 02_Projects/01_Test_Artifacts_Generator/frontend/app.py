import streamlit as st
import requests
import pandas as pd
import io

st.set_page_config(page_title="TestCraft AI", layout="wide")
st.title("🧪 TestCraft AI - Test Artifacts Generator")

backend_generate = "http://127.0.0.1:8000/generate"

requirement = st.text_area("Enter Requirement / User Story", height=200)

if "generated" not in st.session_state:
    st.session_state.generated = None

col1, col2 = st.columns(2)

with col1:
    if st.button("Generate 🚀"):
        if requirement.strip() == "":
            st.warning("Please enter requirement")
        else:
            with st.spinner("Generating artifacts..."):
                resp = requests.post(backend_generate, json={"requirement": requirement})

            if resp.status_code != 200:
                st.error("Backend error")
                st.write(resp.text)
            else:
                st.session_state.generated = resp.json()

with col2:
    if st.button("Clear 🧹"):
        st.session_state.generated = None
        st.rerun()

st.divider()

data = st.session_state.generated

if data:
    scenarios = data.get("scenarios", [])
    test_cases = data.get("test_cases", [])
    test_data = data.get("test_data", [])
    bdd = data.get("bdd", {})

    tab1, tab2, tab3, tab4 = st.tabs(["✅ Scenarios", "🧪 Test Cases", "📌 BDD", "🧾 Test Data"])

    with tab1:
        st.dataframe(pd.DataFrame(scenarios), use_container_width=True)

    with tab2:
        tc_rows = []
        for tc in test_cases:
            row = dict(tc)
            steps = row.get("steps", [])
            if isinstance(steps, list):
                row["steps"] = "\n".join([f"{i+1}. {s}" for i, s in enumerate(steps)])
            tc_rows.append(row)
        st.dataframe(pd.DataFrame(tc_rows), use_container_width=True)

         # Excel Download
        df_tc = pd.DataFrame(tc_rows)

        excel_buffer = io.BytesIO()
        with pd.ExcelWriter(excel_buffer, engine="openpyxl") as writer:
            df_tc.to_excel(writer, index=False, sheet_name="TestCases")
        
            ws = writer.sheets["TestCases"]
                    # Auto column width + wrap text
            from openpyxl.styles import Alignment, Font
            from openpyxl.utils import get_column_letter

            for col_idx, col_name in enumerate(df_tc.columns, start=1):
                col_letter = get_column_letter(col_idx)

                # set width based on max length
                max_len = max(
                    df_tc[col_name].astype(str).map(len).max(),
                    len(col_name)
                )
                ws.column_dimensions[col_letter].width = min(max_len + 5, 50)

                # wrap all cells
                for row in range(1, len(df_tc) + 2):
                    ws[f"{col_letter}{row}"].alignment = Alignment(wrap_text=True, vertical="top")

            # header style
            for cell in ws[1]:
                cell.font = Font(bold=True)

        st.download_button(
            label="⬇️ Download Test Cases Excel",
            data=excel_buffer.getvalue(),
            file_name="test_cases.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )


    with tab3:
        feature_name = bdd.get("feature_name", "")
        description = bdd.get("description", "")
        scenarios_list = bdd.get("scenarios", [])

        feature_text = f"Feature: {feature_name}\n{description}\n\n"
        for sc in scenarios_list:
            feature_text += sc + "\n\n"

        st.code(feature_text.strip(), language="gherkin")

        st.download_button(
            label="⬇️ Download .feature File",
            data=feature_text.strip(),
            file_name="login.feature",
            mime="text/plain"
        )

    with tab4:
        st.dataframe(pd.DataFrame(test_data), use_container_width=True)

else:
    st.info("Enter requirement and click Generate 🚀")
